import re
import os
import openpyxl
from datetime import datetime

#in $$
brackets_regex = re.compile(r'\[(.*?)\]') # [$2022-06-16 14:56:07.687$] INFO [$42432$] [$4320$] [$2_10_297_770$] Process 2_10_297_770 tile
tile_name_regex = re.compile(r'\d+\w\d+\w\d+\w\d+') # [2022-06-20 17:38:02.632] INFO [18560] [8224] [$0_2_0_0$] Start process tile $0_2_0_0$
start_chart_name_regex = re.compile(r'Start process chart (.*?)$') # [2022-06-16 14:56:07.833] INFO [42432] [4320] [2_10_297_770] Start Clip $CA579159$ chart...
finish_chart_name_regex = re.compile(r'Finish process chart (.*?)$') # [2022-06-20 17:38:02.711] INFO [18560] [8224] [0_2_0_0] Finish process chart &2T1AIRWS_0&
is_operation_end_regex = re.compile(r'\w+\stook (.*?) sec') # [2022-06-16 14:56:07.852] DEBUG [42432] [4320] [2_10_297_770] Clipping took $0.0186$ sec
operation_name_regex = re.compile(r'\w+') # [2022-06-16 14:56:07.852] DEBUG [42432] [4320] [2_10_297_770] #Clipping# took $0.0186$ sec

def convert_date_time_to_float(time:datetime)-> float:
    return float(str(time.seconds) + '.' + str(time.microseconds // 1000))

class ItemData():
    def __init__(self, name:str = None, time: datetime = None):
        self.name = name
        self.time = time

    def is_empty(self):
        return self.name is None and self.time is None

class TileStats():
    def __init__(self, name, proccessing_time, charts_n, operations):
        self.__name: str  = name
        self.__processing_time: float = proccessing_time
        self.__charts_n: int = charts_n
        self.__operations: dict = operations
    
    @property
    def processing_time(self) -> float:
        return self.__processing_time

    def get_operations(self) -> int:
        return self.__operations.keys()

    def get_sheet_view(self) -> list:
        sheet_view = [self.__name, self.__processing_time, self.__charts_n]

        for item in self.__operations.items():
            sheet_view.append(item[1])

        return sheet_view
            

class Tile():
    def __init__(self, name: str, start_time: datetime):
        self.__name = name
        self.__start_time = start_time
        self.__finish_time: datetime
        self.__operations: dict = {}
        
    def set_last_message_time(self, time: datetime) -> None:
        self.__finish_time = time

    def add_operation(self, name:str, time:str) -> None:
        self.__operations.setdefault(name, []).append(time)

    def finish(self) -> TileStats:
        tile_operations = {}
        for item in self.__operations.items():
            tile_operations[item[0]] = sum([float(i) for i in item[1]])

        stats = TileStats(self.__name,
                          convert_date_time_to_float(self.__finish_time - self.__start_time),
                          len(self.__operations['Clipping']),
                          tile_operations)

        return stats

class ChartStats():
    def __init__(self, name, count, sum):
        self.__name: str = name
        self.__count: int = count
        self.__sum: float = sum

    @property
    def sum(self) -> float:
        return self.__sum

    def get_sheet_view(self) -> list:
        return [self.__name, self.__count, self.__sum]

class Chart():
    def __init__(self, name: str):
        self.__name = name
        self.__tiles: dict = {}
        
    def add_start_time(self, tile_name:str, time:datetime):
        self.__tiles[tile_name] = [time]

    def add_finish_time(self, tile_name:str, time:datetime):
        self.__tiles[tile_name].append(time)

    def finish(self) -> ChartStats:
        summa = 0
        for item in self.__tiles.items():
            summa += convert_date_time_to_float(item[1][1] - item[1][0])

        stats = ChartStats(self.__name,
                           len(self.__tiles), 
                           summa)

        return stats

def collect_log_statistic(log, tiles:dict, charts:dict): # log: _io.TextIOWrapper
    for line in log:
        if is_message_from_tile(line):
            tile_data: ItemData = try_parse_tile(line)
            if not tile_data.is_empty():
                tiles[tile_data.name] = Tile(tile_data.name, tile_data.time)

            chart_start_data: ItemData = try_start_parse_chart(line) 
            if not chart_start_data.is_empty():
                charts.setdefault(chart_start_data.name, Chart(chart_start_data.name)).add_start_time(find_tile_name(line), chart_start_data.time)

            chart_finish_data: ItemData = try_finish_parse_chart(line) 
            if not chart_finish_data.is_empty():
                charts[chart_finish_data.name].add_finish_time(find_tile_name(line), chart_finish_data.time)

            if is_operation_end(line):
                active_tile = find_tile_name(line)
                tiles[active_tile].set_last_message_time(find_date(line))
                message = find_operation_end_message(line)
                tiles[active_tile].add_operation(find_operation_name(message), find_operation_time(message))

def is_message_from_tile(s:str) -> bool:
    if tile_name_regex.search(s) is not None:
        return True
    return False

def try_parse_tile(s:str) -> ItemData:
    if 'Start process tile' in s:
        return ItemData(find_tile_name(s), find_date(s))

    return ItemData()

def find_tile_name(s:str) -> str:
    return tile_name_regex.search(s).group(0)

def try_start_parse_chart(s:str) -> ItemData:
    if start_chart_name_regex.search(s) is not None:
        return ItemData(start_chart_name_regex.search(s).group(1), find_date(s))

    return ItemData()

def try_finish_parse_chart(s:str) -> ItemData:
    if finish_chart_name_regex.search(s) is not None:
        return ItemData(finish_chart_name_regex.search(s).group(1), find_date(s))

    return ItemData()

def find_date(s:str) -> datetime:
    return datetime.fromisoformat(brackets_regex.findall(s)[0])

def is_operation_end(s:str) -> bool:
    if is_operation_end_regex.search(s) is not None:
        return True
    return False

def find_operation_end_message(s:str) -> str:
    return is_operation_end_regex.search(s).group(0)

def find_operation_name(s:str) -> str:
    return operation_name_regex.search(s).group(0)

def find_operation_time(s:str) -> str:
    return is_operation_end_regex.search(s).group(1)

def create_tiles_sheet(book, tiles: list):
    book.create_sheet("Tiles")
    sheet : openpyxl.worksheet = book.worksheets[0]

    sheet.insert_rows(0)
    sheet["A1"].value = "Name"
    sheet.column_dimensions["A"].width = 50
    sheet["B1"].value = "Time"
    sheet.column_dimensions["B"].width = 50
    sheet["C1"].value = "Charts"
    sheet.column_dimensions["C"].width = 50

    ascii_i = 68 #D
    for operation in tiles[0].get_operations():
        current_letter = str(chr(ascii_i))
        sheet[current_letter + '1'].value = operation
        sheet.column_dimensions[current_letter].width = 50
        ascii_i += 1

    for tile in tiles:
        sheet.append(tile.get_sheet_view())

    sheet.auto_filter.ref = "A1:" + current_letter + str(len(tiles))

def create_charts_sheet(book, charts: list):
    book.create_sheet("Charts")
    sheet : openpyxl.worksheet = book.worksheets[1]

    sheet.insert_rows(0)
    sheet["A1"].value = "Name"
    sheet.column_dimensions["A"].width = 50
    sheet["B1"].value = "Time"
    sheet.column_dimensions["B"].width = 50
    sheet["C1"].value = "Tiles"
    sheet.column_dimensions["C"].width = 50

    for chart in charts:
        sheet.append(chart.get_sheet_view())

    sheet.auto_filter.ref = "A1:C" + str(len(charts))

def create_stats_sheet(book, tiles: list, charts: list):
    book.create_sheet("Stats")
    sheet : openpyxl.worksheet = book.worksheets[2]

    tiles = sorted(tiles, key=lambda elem: elem.processing_time)
    charts = sorted(charts, key=lambda elem: elem.sum)

    tilesN = len(tiles)
    chartsN = len(charts)

    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 50

    sheet.append(["Tiles number", tilesN])
    sheet.append(["Charts number", chartsN])

    sheet.append(["Tiles median", tiles[tilesN // 2].processing_time])
    sheet.append(["Chart median", charts[chartsN // 2].sum])

    sheet.append(["Tiles quantile up", tiles[int(tilesN * 0.1)].processing_time])
    sheet.append(["Charts quantile up", charts[int(chartsN * 0.1)].sum])

    sheet.append(["Tiles quantile down", tiles[int(tilesN - tilesN * 0.1 - 1)].processing_time])
    sheet.append(["Charts quantile down", charts[int(chartsN - chartsN * 0.1 - 1)].sum])

def create_excel_view(tiles: dict, charts: dict):
    book = openpyxl.Workbook()
    book.remove(book.active)
    
    tiles_stats = []
    chart_stats = []

    for tile in tiles.items():
        tiles_stats.append(tiles[tile[0]].finish())

    for chart in charts.items():
        chart_stats.append(charts[chart[0]].finish())

    create_tiles_sheet(book, tiles_stats)
    create_charts_sheet(book, chart_stats)
    create_stats_sheet(book, tiles_stats, chart_stats)

    book.save("LogStats.xlsx")

if __name__ == "__main__":
    for root, dirs, files in os.walk(r"Logs"):
        for file in files:
            fileName = os.path.join(root, file)
            print(fileName)

            tiles = {}
            charts = {}
            with open(fileName, 'r') as log:
                collect_log_statistic(log, tiles, charts)

    print("Stats collected succesfully")

    create_excel_view(tiles, charts)

    print("View created successfully")
